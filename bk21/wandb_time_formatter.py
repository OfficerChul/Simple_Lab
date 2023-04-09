import openpyxl

filename = 'BK21-result.xlsx'

wb_obj = openpyxl.load_workbook(filename)
print(wb_obj.sheetnames)
start_idx = [6, 20, 34]
mix_start_idx = [6, 17, 28]

sheet_list = ['a', 'b', 'c']
mix_sheet_list = ['mix-a', 'mix-b', 'mix-c']

for sheet in mix_sheet_list:
    sheet = wb_obj[sheet]
    for si in mix_start_idx:
        for i in range(9):
            cell_obj = sheet.cell(row=i + si, column=4)
            val = cell_obj.value

            h, m, s = 0, 0, 0
            print(f'val: {val}')
            time = val.split(" ")
            if len(time) == 3:
                h = int(time[0][0: time[0].find('h')])
                m = int(time[1][0: time[1].find('m')])
                s = int(time[2][0: time[2].find('s')])
            else:
                m = int(time[0][0: time[0].find('m')])
                s = int(time[1][0: time[1].find('s')])

            new_val = 3600 * h + 60 * m + s
            print(f'{val} -> {new_val}')
            cell_obj.value = new_val

wb_obj.save(filename)
        


